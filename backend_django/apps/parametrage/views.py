from django.db import connection
from rest_framework import generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from .models import (
    Nationalite, FormeJuridique, DomaineActivite, Fonction,
    TypeDocument, TypeDemande, Localite, Tarif, Signataire,
)
from .serializers import (
    NationaliteSerializer, FormeJuridiqueSerializer, DomaineActiviteSerializer,
    FonctionSerializer, TypeDocumentSerializer, TypeDemandeSerializer,
    LocaliteSerializer, TarifSerializer, SignataireSerializer,
)
from apps.core.permissions import LectureAgentModifGreffier, EstGreffier


# ── Nationalités ──────────────────────────────────────────────────────────────
class NationaliteListCreate(generics.ListCreateAPIView):
    """Lecture : tout le personnel. Écriture : greffier (CDC §3.2)."""
    permission_classes = [LectureAgentModifGreffier]
    queryset         = Nationalite.objects.filter(actif=True)
    serializer_class = NationaliteSerializer

class NationaliteDetail(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [LectureAgentModifGreffier]
    queryset         = Nationalite.objects.all()
    serializer_class = NationaliteSerializer


class NationaliteImportView(APIView):
    """
    POST /parametrage/nationalites/import/
    Importe des nationalités depuis un fichier Excel (.xlsx).
    Colonnes requises : code, libelle_fr, libelle_ar
    Réservé au greffier (EstGreffier).
    Retourne : { total, created, duplicates, errors }
    """
    permission_classes = [EstGreffier]
    parser_classes     = [MultiPartParser]

    REQUIRED_COLS = {'code', 'libelle_fr', 'libelle_ar'}

    def post(self, request):
        try:
            import openpyxl
        except ImportError:
            return Response(
                {'detail': 'openpyxl non disponible sur le serveur.'},
                status=500,
            )

        fichier = request.FILES.get('fichier')
        if not fichier:
            return Response({'detail': 'Aucun fichier fourni (champ : fichier).'}, status=400)

        # ── Lecture du classeur ───────────────────────────────────────────────
        try:
            wb = openpyxl.load_workbook(fichier, read_only=True, data_only=True)
        except Exception as exc:
            return Response({'detail': f'Fichier Excel invalide : {exc}'}, status=400)

        ws = wb.active

        # ── Extraction des en-têtes (1re ligne) ───────────────────────────────
        headers = [str(cell.value).strip().lower() if cell.value is not None else ''
                   for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        missing = self.REQUIRED_COLS - set(headers)
        if missing:
            return Response(
                {'detail': f'Colonnes manquantes dans le fichier : {", ".join(sorted(missing))}'},
                status=400,
            )

        col_idx = {name: headers.index(name) for name in self.REQUIRED_COLS}

        # ── Parcours des lignes ───────────────────────────────────────────────
        total      = 0
        created    = 0
        duplicates = []
        errors     = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Ignorer les lignes entièrement vides
            if all(v is None or str(v).strip() == '' for v in row):
                continue

            total += 1

            def cell(col):
                v = row[col_idx[col]]
                return str(v).strip() if v is not None else ''

            code       = cell('code')
            libelle_fr = cell('libelle_fr')
            libelle_ar = cell('libelle_ar')

            # ── Validation : champs obligatoires ─────────────────────────────
            manquants = []
            if not code:       manquants.append('code')
            if not libelle_fr: manquants.append('libelle_fr')
            if not libelle_ar: manquants.append('libelle_ar')
            if manquants:
                errors.append({
                    'ligne': row_num,
                    'code': code or '(vide)',
                    'raison': f'Champ(s) obligatoire(s) absent(s) : {", ".join(manquants)}',
                })
                continue

            # ── Tronquer si dépassement max_length ────────────────────────────
            code = code[:5]

            # ── Détection doublon ─────────────────────────────────────────────
            if Nationalite.objects.filter(code=code).exists():
                duplicates.append({'ligne': row_num, 'code': code})
                continue

            # ── Création ──────────────────────────────────────────────────────
            try:
                Nationalite.objects.create(
                    code=code,
                    libelle_fr=libelle_fr[:100],
                    libelle_ar=libelle_ar[:100],
                    actif=True,
                )
                created += 1
            except Exception as exc:
                errors.append({
                    'ligne': row_num,
                    'code': code,
                    'raison': str(exc),
                })

        wb.close()

        return Response({
            'total':      total,
            'created':    created,
            'duplicates': duplicates,
            'errors':     errors,
        }, status=201 if created > 0 else 200)


# ── Formes juridiques ─────────────────────────────────────────────────────────
class FormeJuridiqueListCreate(generics.ListCreateAPIView):
    permission_classes = [LectureAgentModifGreffier]
    queryset         = FormeJuridique.objects.filter(actif=True)
    serializer_class = FormeJuridiqueSerializer

class FormeJuridiqueDetail(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [LectureAgentModifGreffier]
    queryset         = FormeJuridique.objects.all()
    serializer_class = FormeJuridiqueSerializer


# ── Domaines d'activité ───────────────────────────────────────────────────────
class DomaineActiviteListCreate(generics.ListCreateAPIView):
    permission_classes = [LectureAgentModifGreffier]
    queryset         = DomaineActivite.objects.filter(actif=True)
    serializer_class = DomaineActiviteSerializer

class DomaineActiviteDetail(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [LectureAgentModifGreffier]
    queryset         = DomaineActivite.objects.all()
    serializer_class = DomaineActiviteSerializer


# ── Fonctions ─────────────────────────────────────────────────────────────────
class FonctionListCreate(generics.ListCreateAPIView):
    permission_classes = [LectureAgentModifGreffier]
    queryset         = Fonction.objects.filter(actif=True)
    serializer_class = FonctionSerializer

class FonctionDetail(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [LectureAgentModifGreffier]
    queryset         = Fonction.objects.all()
    serializer_class = FonctionSerializer


# ── Types de documents ────────────────────────────────────────────────────────
class TypeDocumentListCreate(generics.ListCreateAPIView):
    permission_classes = [LectureAgentModifGreffier]
    queryset         = TypeDocument.objects.filter(actif=True)
    serializer_class = TypeDocumentSerializer

class TypeDocumentDetail(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [LectureAgentModifGreffier]
    queryset         = TypeDocument.objects.all()
    serializer_class = TypeDocumentSerializer


# ── Types de demandes ─────────────────────────────────────────────────────────
class TypeDemandeListCreate(generics.ListCreateAPIView):
    permission_classes = [LectureAgentModifGreffier]
    queryset         = TypeDemande.objects.filter(actif=True)
    serializer_class = TypeDemandeSerializer

class TypeDemandeDetail(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [LectureAgentModifGreffier]
    queryset         = TypeDemande.objects.all()
    serializer_class = TypeDemandeSerializer


# ── Localités ─────────────────────────────────────────────────────────────────
class LocaliteListCreate(generics.ListCreateAPIView):
    permission_classes = [LectureAgentModifGreffier]
    serializer_class = LocaliteSerializer

    def get_queryset(self):
        qs = Localite.objects.filter(actif=True)
        type_loc = self.request.query_params.get('type')
        if type_loc:
            qs = qs.filter(type=type_loc)
        return qs

class LocaliteDetail(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [LectureAgentModifGreffier]
    queryset         = Localite.objects.all()
    serializer_class = LocaliteSerializer


# ── Tarifs — administration pure, greffier uniquement ─────────────────────────
class TarifListCreate(generics.ListCreateAPIView):
    """CDC §3.3 : administration des tarifs réservée au greffier."""
    permission_classes = [EstGreffier]
    queryset         = Tarif.objects.filter(actif=True)
    serializer_class = TarifSerializer

class TarifDetail(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [EstGreffier]
    queryset         = Tarif.objects.all()
    serializer_class = TarifSerializer


# ── Signataires — administration pure, greffier uniquement ───────────────────
class SignataireListCreate(generics.ListCreateAPIView):
    """CDC §3.3 : administration des signataires réservée au greffier."""
    permission_classes = [EstGreffier]
    queryset         = Signataire.objects.all()
    serializer_class = SignataireSerializer

class SignataireDetail(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [EstGreffier]
    queryset         = Signataire.objects.all()
    serializer_class = SignataireSerializer


# ── Numérotation (accès direct à sequences_numerotation) ─────────────────────

class NumerotationListView(APIView):
    """Liste toutes les séquences de numérotation. Réservé au greffier (CDC §3.3)."""
    permission_classes = [EstGreffier]

    def get(self, request):
        with connection.cursor() as c:
            c.execute("""
                SELECT code, prefixe, annee, dernier_num, nb_chiffres, updated_at
                FROM sequences_numerotation
                ORDER BY code
            """)
            rows = c.fetchall()
        labels = {
            # Numérotation analytique — deux séquences distinctes (règle de parité RCCM)
            'RA_PP': 'N° Analytique Personne Physique (PAIR, continu — ex : 000002, 000004…)',
            'RA_PM': 'N° Analytique Personne Morale / Succursale (IMPAIR, continu — ex : 000001, 000003…)',
            # Ancien compteur unifié — conservé en lecture, ne doit plus être incrémenté
            'RA':    'N° Analytique (ancien compteur unifié — référence historique, lecture seule)',
            # Numérotation chronologique — annuelle (règle RCCM)
            'CHRONO':'N° Chronologique (annuel, remis à 1 chaque 1er janvier)',
            'RC':    'N° Chronologique (ancien code — lecture seule)',
            'DMD':   'N° Demande (annuel)',
            'DEP':   'N° Dépôt (annuel)',
            'MOD':   'N° Modification (annuel)',
            'RAD':   'N° Radiation (annuel)',
            'CES':   'N° Cession (annuel)',
        }
        data = [
            {
                'code':        r[0],
                'libelle':     labels.get(r[0], r[0]),
                'prefixe':     r[1],
                'annee':       r[2],
                'dernier_num': r[3],
                'nb_chiffres': r[4],
                'updated_at':  r[5],
            }
            for r in rows
        ]
        return Response(data)


class NumerotationUpdateView(APIView):
    """
    Modifie manuellement le compteur d'une séquence.
    PUT /parametrage/numerotation/<code>/
    Body : { "dernier_num": <int> }
    Réservé au greffier (CDC §3.3).
    """
    permission_classes = [EstGreffier]

    def put(self, request, code):
        dernier_num = request.data.get('dernier_num')
        if dernier_num is None:
            return Response({'detail': 'Le champ dernier_num est requis.'}, status=400)
        try:
            dernier_num = int(dernier_num)
        except (ValueError, TypeError):
            return Response({'detail': 'dernier_num doit être un entier.'}, status=400)

        # ── Règles de parité RCCM (bloquantes) ────────────────────────────────
        # RA_PP : séquence Personne Physique → doit être PAIR
        if code == 'RA_PP' and dernier_num % 2 != 0:
            return Response(
                {'detail': 'Le N° Analytique PP (RA_PP) doit être PAIR (2, 4, 6…). '
                           f'La valeur {dernier_num} est impaire — opération refusée.'},
                status=400,
            )
        # RA_PM : séquence Personne Morale / Succursale → doit être IMPAIR
        if code == 'RA_PM' and dernier_num % 2 == 0:
            return Response(
                {'detail': 'Le N° Analytique PM/SC (RA_PM) doit être IMPAIR (1, 3, 5…). '
                           f'La valeur {dernier_num} est paire — opération refusée.'},
                status=400,
            )
        # RA (ancien compteur) : conserve la règle impair pour rétrocompatibilité
        if code == 'RA' and dernier_num % 2 == 0:
            return Response({'detail': 'Le N° Analytique (RA) doit être impair (1, 3, 5…).'}, status=400)

        with connection.cursor() as c:
            c.execute("""
                UPDATE sequences_numerotation
                SET dernier_num = %s, updated_at = NOW()
                WHERE code = %s
            """, [dernier_num, code])
            updated = c.rowcount

        if updated == 0:
            return Response({'detail': f'Code « {code} » introuvable.'}, status=404)

        return Response({'message': f'Numérotation « {code} » mise à jour → {dernier_num}.'})
